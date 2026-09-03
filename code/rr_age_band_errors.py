# -*- coding: utf-8 -*-
"""부록 표 12-13 재현(패키지 루트에서 실행): 2024 연령대(7개 밴드, 성별 합산)별 부호오차(합성 - 실측, %p),
모델 2종(Gemini, EXAONE) × 이진 8지표.

실측: 2024 만 10세 이상 표본(8,675명)의 밴드별 가중 비율(WT, 두 성별 합산).
합성: 밴드 내 합성 페르소나의 단순평균(비가중 합산; 두 성별 셀은 10대 478/490, 그 밖의 밴드 600/600).
    이 규칙이 표 12-13 의 정수 112개를 모두 재현한다(부호오차%p).
대안 규칙(사후층화: 밴드 내 두 성별 셀 평균을 해당 지표 유효 응답의 실측 가중 셀 점유율로 가중; 셀 오차의 실측 점유율
가중평균과 동일)은 부호오차_사후층화%p 로 병기하며 4개 셀에서 정수 반올림이 1 다르다(콘솔에 두 규칙을 함께 출력).

시트 심사_연령대별_부호오차 만 갱신한다: openpyxl 로 해당 시트만 제거 후 다시 쓰고,
다른 시트의 셀 값 해시를 전후로 비교하여 변경이 없음을 확인한다."""
import hashlib, sys
import numpy as np, pandas as pd
from openpyxl import load_workbook
from rr_common import BIN, AGES, SYN_FILES, OUT_XLSX, load_real, load_syn

SHEET = "심사_연령대별_부호오차"
LABEL = {"AI_이용여부": "AI use", "OTT_이용": "OTT", "유튜브_이용": "YouTube", "숏폼_이용": "Short-form",
         "SNS_이용": "SNS", "메신저_이용": "Messenger", "메타버스_이용": "Metaverse", "콘텐츠구독_이용": "Subscription"}


def band_values(real, syn, v):
    """밴드별 (실측 가중률, 합성 사후층화율, 합성 비가중 평균) 반환."""
    r = real[[v, "WT", "성별", "연령대"]].dropna()
    out = {}
    for g in AGES:
        rg = r[r["연령대"] == g]
        act = np.average(rg[v], weights=rg["WT"])
        share = rg.groupby("성별")["WT"].sum(); share = share / share.sum()
        sg = syn[syn["연령대"] == g][[v, "성별"]].dropna()
        cm = sg.groupby("성별")[v].mean()
        common = [s for s in cm.index if s in share.index]
        ps = float(sum(share[s] * cm[s] for s in common) / sum(share[s] for s in common))
        out[g] = (float(act), ps, float(sg[v].mean()))
    return out


def rint(x):
    """0.5 는 0 에서 먼 쪽으로 반올림(표기 정수와 비교용)."""
    return int(np.sign(x) * np.floor(abs(x) + 0.5))


def sheet_hashes(path):
    """시트별 셀 값 해시(행 끝의 빈 셀은 제거하여 저장 방식 차이에 둔감하게 한다)."""
    wb = load_workbook(path, read_only=True, data_only=True); h = {}
    for ws in wb.worksheets:
        m = hashlib.sha256()
        for row in ws.iter_rows(values_only=True):
            row = list(row)
            while row and row[-1] is None: row.pop()
            if row: m.update(repr(row).encode("utf-8"))
        h[ws.title] = m.hexdigest()
    wb.close(); return h


real = load_real(2024, with_hid=False)
print(f"실측 2024 만 10세 이상 n={len(real)}")
rows = []; console = {}
for m, f in SYN_FILES.items():
    syn = load_syn(f)
    for v in BIN:
        bv = band_values(real, syn, v)
        rows.append({"모델": m, "변수": v, "값종류": "실측_가중률%", **{g: round(bv[g][0] * 100, 2) for g in AGES}})
        rows.append({"모델": m, "변수": v, "값종류": "합성_비가중률%", **{g: round(bv[g][2] * 100, 2) for g in AGES}})
        rows.append({"모델": m, "변수": v, "값종류": "부호오차%p", **{g: round((bv[g][2] - bv[g][0]) * 100, 2) for g in AGES}})
        rows.append({"모델": m, "변수": v, "값종류": "합성_사후층화율%", **{g: round(bv[g][1] * 100, 2) for g in AGES}})
        rows.append({"모델": m, "변수": v, "값종류": "부호오차_사후층화%p", **{g: round((bv[g][1] - bv[g][0]) * 100, 2) for g in AGES}})
        console[(m, v)] = ([rint((bv[g][2] - bv[g][0]) * 100) for g in AGES], [rint((bv[g][1] - bv[g][0]) * 100) for g in AGES])
df = pd.DataFrame(rows)

for m in SYN_FILES:
    print(f"\n[{m}] 부호오차(정수 반올림): 비가중 합산 규칙(표 12-13) | 사후층화 규칙")
    for v in BIN:
        a, b = console[(m, v)]
        print(f"  {LABEL[v]:<13}" + " ".join(f"{x:+4d}" for x in a) + "  |  " + " ".join(f"{x:+4d}" for x in b))

before = sheet_hashes(OUT_XLSX)
wb = load_workbook(OUT_XLSX)
if SHEET in wb.sheetnames: del wb[SHEET]
ws = wb.create_sheet(SHEET)
ws.append(list(df.columns))
for r in df.itertuples(index=False):
    ws.append([x.item() if hasattr(x, "item") else x for x in r])
wb.save(OUT_XLSX)
after = sheet_hashes(OUT_XLSX)
changed = [s for s in before if s != SHEET and before[s] != after.get(s)]
print(f"\n워크북 시트 갱신: {SHEET} | 시트 수 {len(before)} -> {len(after)} | 다른 시트 변경: {changed if changed else '없음'}")
if changed: sys.exit(1)
